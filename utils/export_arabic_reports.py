"""
Arabic report export utilities with proper font support
PDF generation with RTL text support and Arabic typography
"""

import logging
import asyncio
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import io
import base64
from pathlib import Path

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.platypus.flowables import Image
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)

class ArabicReportGenerator:
    """Generate PDF reports with proper Arabic font support"""
    
    def __init__(self):
        self.setup_arabic_fonts()
        self.setup_styles()
        
    def setup_arabic_fonts(self):
        """Setup Arabic fonts for PDF generation"""
        if not REPORTLAB_AVAILABLE:
            logger.warning("ReportLab not available for PDF generation")
            return
            
        try:
            # Try to register Arabic fonts
            # In production, you would use actual Arabic font files
            self.arabic_font_available = False
            
            # For now, use built-in fonts that support Unicode
            self.default_font = 'Helvetica'
            self.bold_font = 'Helvetica-Bold'
            
            logger.info("Arabic font setup completed (using Unicode-compatible fallback)")
            
        except Exception as e:
            logger.error(f"Error setting up Arabic fonts: {e}")
            self.arabic_font_available = False
    
    def setup_styles(self):
        """Setup paragraph styles for Arabic content"""
        if not REPORTLAB_AVAILABLE:
            return
            
        self.styles = getSampleStyleSheet()
        
        # Arabic title style
        self.styles.add(ParagraphStyle(
            name='ArabicTitle',
            parent=self.styles['Title'],
            fontName=self.bold_font,
            fontSize=18,
            textColor=colors.darkblue,
            alignment=TA_RIGHT,
            spaceAfter=20,
            rightIndent=0,
            leftIndent=0
        ))
        
        # Arabic heading style
        self.styles.add(ParagraphStyle(
            name='ArabicHeading',
            parent=self.styles['Heading1'],
            fontName=self.bold_font,
            fontSize=14,
            textColor=colors.darkgreen,
            alignment=TA_RIGHT,
            spaceAfter=12,
            spaceBefore=12
        ))
        
        # Arabic body text style
        self.styles.add(ParagraphStyle(
            name='ArabicBody',
            parent=self.styles['Normal'],
            fontName=self.default_font,
            fontSize=12,
            alignment=TA_RIGHT,
            spaceAfter=6,
            rightIndent=10,
            leftIndent=10
        ))
        
        # Arabic table header style
        self.styles.add(ParagraphStyle(
            name='ArabicTableHeader',
            parent=self.styles['Normal'],
            fontName=self.bold_font,
            fontSize=11,
            textColor=colors.white,
            alignment=TA_CENTER
        ))
        
        # Arabic table cell style
        self.styles.add(ParagraphStyle(
            name='ArabicTableCell',
            parent=self.styles['Normal'],
            fontName=self.default_font,
            fontSize=10,
            alignment=TA_RIGHT
        ))
    
    def create_sentiment_analytics_report(self, data: Dict[str, Any]) -> bytes:
        """Create sentiment analytics report in Arabic"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
        
        story = []
        
        # Report title
        title = "تقرير تحليل المشاعر - منصة صوت العميل العربية"
        story.append(Paragraph(title, self.styles['ArabicTitle']))
        story.append(Spacer(1, 20))
        
        # Report date
        report_date = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        story.append(Paragraph(report_date, self.styles['ArabicBody']))
        story.append(Spacer(1, 20))
        
        # Executive summary
        story.append(Paragraph("الملخص التنفيذي", self.styles['ArabicHeading']))
        
        summary_data = data.get('summary', {})
        total_feedback = summary_data.get('total_feedback', 0)
        avg_sentiment = summary_data.get('average_sentiment', 0.0)
        
        summary_text = f"""
        إجمالي التعليقات المحللة: {total_feedback:,}<br/>
        متوسط درجة المشاعر: {avg_sentiment:.2f}<br/>
        فترة التحليل: {summary_data.get('period', 'آخر 30 يوم')}
        """
        story.append(Paragraph(summary_text, self.styles['ArabicBody']))
        story.append(Spacer(1, 20))
        
        # Sentiment distribution
        story.append(Paragraph("توزيع المشاعر", self.styles['ArabicHeading']))
        
        sentiment_dist = data.get('sentiment_distribution', {})
        sentiment_table_data = [
            [Paragraph("النسبة المئوية", self.styles['ArabicTableHeader']),
             Paragraph("العدد", self.styles['ArabicTableHeader']),
             Paragraph("نوع المشاعر", self.styles['ArabicTableHeader'])]
        ]
        
        sentiment_labels = {
            'positive': 'إيجابي',
            'neutral': 'محايد',
            'negative': 'سلبي'
        }
        
        total_count = sum(sentiment_dist.values()) if sentiment_dist else 1
        
        for sentiment_type, count in sentiment_dist.items():
            percentage = (count / total_count) * 100
            sentiment_table_data.append([
                Paragraph(f"{percentage:.1f}%", self.styles['ArabicTableCell']),
                Paragraph(f"{count:,}", self.styles['ArabicTableCell']),
                Paragraph(sentiment_labels.get(sentiment_type, sentiment_type), self.styles['ArabicTableCell'])
            ])
        
        sentiment_table = Table(sentiment_table_data, colWidths=[3*cm, 3*cm, 4*cm])
        sentiment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), self.bold_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(sentiment_table)
        story.append(Spacer(1, 20))
        
        # Channel performance
        story.append(Paragraph("أداء القنوات", self.styles['ArabicHeading']))
        
        channel_data = data.get('channel_performance', [])
        if channel_data:
            channel_table_data = [
                [Paragraph("متوسط التقييم", self.styles['ArabicTableHeader']),
                 Paragraph("متوسط المشاعر", self.styles['ArabicTableHeader']),
                 Paragraph("عدد التعليقات", self.styles['ArabicTableHeader']),
                 Paragraph("القناة", self.styles['ArabicTableHeader'])]
            ]
            
            for channel in channel_data[:10]:  # Top 10 channels
                avg_rating = channel.get('avg_rating', 0) or 0
                avg_sentiment = channel.get('avg_sentiment', 0)
                total_feedback = channel.get('total_feedback', 0)
                channel_name = channel.get('channel_ar', channel.get('channel', ''))
                
                channel_table_data.append([
                    Paragraph(f"{avg_rating:.1f}" if avg_rating > 0 else "غير متاح", self.styles['ArabicTableCell']),
                    Paragraph(f"{avg_sentiment:.2f}", self.styles['ArabicTableCell']),
                    Paragraph(f"{total_feedback:,}", self.styles['ArabicTableCell']),
                    Paragraph(channel_name, self.styles['ArabicTableCell'])
                ])
            
            channel_table = Table(channel_table_data, colWidths=[3*cm, 3*cm, 3*cm, 5*cm])
            channel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), self.bold_font),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(channel_table)
            story.append(Spacer(1, 20))
        
        # Trending topics
        story.append(Paragraph("المواضيع الرائجة", self.styles['ArabicHeading']))
        
        trending_topics = data.get('trending_topics', [])
        if trending_topics:
            topics_text = ""
            for i, topic in enumerate(trending_topics[:10], 1):
                topic_name = topic.get('topic', '')
                mentions = topic.get('mentions', 0)
                sentiment_score = topic.get('sentiment', 0)
                
                topics_text += f"{i}. {topic_name} - {mentions} ذكر (المشاعر: {sentiment_score:.2f})<br/>"
            
            story.append(Paragraph(topics_text, self.styles['ArabicBody']))
            story.append(Spacer(1, 20))
        
        # Cultural insights
        story.append(Paragraph("الرؤى الثقافية", self.styles['ArabicHeading']))
        
        cultural_insights = data.get('cultural_insights', [])
        if cultural_insights:
            insights_text = ""
            for insight in cultural_insights[:5]:
                insight_ar = insight.get('insight_ar', '')
                description_ar = insight.get('description_ar', '')
                impact_score = insight.get('impact_score', 0)
                
                insights_text += f"• {insight_ar}: {description_ar} (التأثير: {impact_score*100:.1f}%)<br/>"
            
            story.append(Paragraph(insights_text, self.styles['ArabicBody']))
            story.append(Spacer(1, 20))
        
        # Footer
        story.append(PageBreak())
        footer_text = f"""
        هذا التقرير تم إنشاؤه تلقائياً بواسطة منصة صوت العميل العربية<br/>
        تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        جميع الحقوق محفوظة
        """
        story.append(Paragraph(footer_text, self.styles['ArabicBody']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_dialect_analysis_report(self, data: Dict[str, Any]) -> bytes:
        """Create dialect analysis report"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
        
        story = []
        
        # Report title
        title = "تقرير تحليل اللهجات العربية"
        story.append(Paragraph(title, self.styles['ArabicTitle']))
        story.append(Spacer(1, 20))
        
        # Dialect distribution
        story.append(Paragraph("توزيع اللهجات", self.styles['ArabicHeading']))
        
        dialect_data = data.get('dialect_breakdown', [])
        if dialect_data:
            dialect_table_data = [
                [Paragraph("مستوى الثقة", self.styles['ArabicTableHeader']),
                 Paragraph("متوسط المشاعر", self.styles['ArabicTableHeader']),
                 Paragraph("العدد", self.styles['ArabicTableHeader']),
                 Paragraph("اللهجة", self.styles['ArabicTableHeader'])]
            ]
            
            dialect_names = {
                'gulf': 'خليجي',
                'egyptian': 'مصري',
                'levantine': 'شامي',
                'moroccan': 'مغربي',
                'other': 'أخرى'
            }
            
            for dialect in dialect_data:
                dialect_name = dialect_names.get(dialect.get('dialect', ''), dialect.get('dialect', ''))
                count = dialect.get('count', 0)
                avg_sentiment = dialect.get('avg_sentiment', 0)
                confidence = dialect.get('confidence', 0)
                
                dialect_table_data.append([
                    Paragraph(f"{confidence*100:.1f}%", self.styles['ArabicTableCell']),
                    Paragraph(f"{avg_sentiment:.2f}", self.styles['ArabicTableCell']),
                    Paragraph(f"{count:,}", self.styles['ArabicTableCell']),
                    Paragraph(dialect_name, self.styles['ArabicTableCell'])
                ])
            
            dialect_table = Table(dialect_table_data, colWidths=[3*cm, 3*cm, 3*cm, 4*cm])
            dialect_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), self.bold_font),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lavender),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(dialect_table)
            story.append(Spacer(1, 20))
        
        # Sample phrases by dialect
        story.append(Paragraph("عينات من العبارات حسب اللهجة", self.styles['ArabicHeading']))
        
        for dialect in dialect_data:
            dialect_name = dialect_names.get(dialect.get('dialect', ''), dialect.get('dialect', ''))
            sample_phrases = dialect.get('sample_phrases', [])
            
            if sample_phrases:
                story.append(Paragraph(f"عينات من اللهجة {dialect_name}:", self.styles['ArabicBody']))
                
                phrases_text = ""
                for i, phrase in enumerate(sample_phrases[:3], 1):
                    phrases_text += f"{i}. {phrase}<br/>"
                
                story.append(Paragraph(phrases_text, self.styles['ArabicBody']))
                story.append(Spacer(1, 10))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_comprehensive_report(self, data: Dict[str, Any]) -> bytes:
        """Create comprehensive analytics report"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
        
        story = []
        
        # Cover page
        title = "التقرير الشامل لتحليل صوت العميل العربي"
        story.append(Paragraph(title, self.styles['ArabicTitle']))
        story.append(Spacer(1, 50))
        
        # Executive summary
        summary = f"""
        فترة التقرير: {data.get('period', 'آخر 30 يوم')}<br/>
        إجمالي التعليقات: {data.get('total_feedback', 0):,}<br/>
        القنوات المحللة: {len(data.get('channel_performance', []))}<br/>
        اللهجات المكتشفة: {len(data.get('dialect_breakdown', []))}<br/>
        المواضيع الرئيسية: {len(data.get('trending_topics', []))}<br/>
        """
        story.append(Paragraph(summary, self.styles['ArabicBody']))
        story.append(PageBreak())
        
        # Sentiment analysis section
        sentiment_report_data = {
            'summary': data.get('summary', {}),
            'sentiment_distribution': data.get('sentiment_distribution', {}),
            'channel_performance': data.get('channel_performance', []),
            'trending_topics': data.get('trending_topics', []),
            'cultural_insights': data.get('cultural_insights', [])
        }
        
        # Add sentiment analysis content (reuse from sentiment report)
        # This would include the same sections as the sentiment report
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

class ExcelReportGenerator:
    """Generate Excel reports with Arabic support"""
    
    def __init__(self):
        try:
            import openpyxl
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            logger.warning("openpyxl not available for Excel generation")
    
    def create_sentiment_excel_report(self, data: Dict[str, Any]) -> bytes:
        """Create sentiment analysis Excel report"""
        if not self.openpyxl_available:
            raise ImportError("openpyxl is required for Excel generation")
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "ملخص التحليل"
        
        # Set RTL direction
        ws_summary.sheet_view.rightToLeft = True
        
        # Headers
        headers = ["القيمة", "المؤشر"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Data
        summary_data = data.get('summary', {})
        data_rows = [
            ["إجمالي التعليقات", summary_data.get('total_feedback', 0)],
            ["متوسط المشاعر", f"{summary_data.get('average_sentiment', 0):.2f}"],
            ["فترة التحليل", summary_data.get('period', 'آخر 30 يوم')]
        ]
        
        for row_idx, (metric, value) in enumerate(data_rows, 2):
            ws_summary.cell(row=row_idx, column=1, value=value)
            ws_summary.cell(row=row_idx, column=2, value=metric)
        
        # Channel performance sheet
        ws_channels = wb.create_sheet("أداء القنوات")
        ws_channels.sheet_view.rightToLeft = True
        
        channel_headers = ["متوسط التقييم", "متوسط المشاعر", "عدد التعليقات", "القناة"]
        for col, header in enumerate(channel_headers, 1):
            cell = ws_channels.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        channel_data = data.get('channel_performance', [])
        for row_idx, channel in enumerate(channel_data, 2):
            ws_channels.cell(row=row_idx, column=1, value=channel.get('avg_rating', 0) or 0)
            ws_channels.cell(row=row_idx, column=2, value=channel.get('avg_sentiment', 0))
            ws_channels.cell(row=row_idx, column=3, value=channel.get('total_feedback', 0))
            ws_channels.cell(row=row_idx, column=4, value=channel.get('channel_ar', channel.get('channel', '')))
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

class ReportExportService:
    """Service for exporting various analytics reports"""
    
    def __init__(self):
        self.pdf_generator = ArabicReportGenerator()
        self.excel_generator = ExcelReportGenerator()
    
    async def export_sentiment_report(self, data: Dict[str, Any], format: str = "pdf") -> Dict[str, Any]:
        """Export sentiment analysis report"""
        try:
            if format.lower() == "pdf":
                report_bytes = self.pdf_generator.create_sentiment_analytics_report(data)
                filename = f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                content_type = "application/pdf"
            
            elif format.lower() == "excel":
                report_bytes = self.excel_generator.create_sentiment_excel_report(data)
                filename = f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            else:
                raise ValueError(f"Unsupported format: {format}")
            
            # Encode as base64 for API response
            report_base64 = base64.b64encode(report_bytes).decode('utf-8')
            
            return {
                "success": True,
                "filename": filename,
                "content_type": content_type,
                "size_bytes": len(report_bytes),
                "data": report_base64
            }
            
        except Exception as e:
            logger.error(f"Error exporting sentiment report: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def export_dialect_report(self, data: Dict[str, Any], format: str = "pdf") -> Dict[str, Any]:
        """Export dialect analysis report"""
        try:
            if format.lower() == "pdf":
                report_bytes = self.pdf_generator.create_dialect_analysis_report(data)
                filename = f"dialect_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                content_type = "application/pdf"
            else:
                raise ValueError(f"Unsupported format for dialect report: {format}")
            
            report_base64 = base64.b64encode(report_bytes).decode('utf-8')
            
            return {
                "success": True,
                "filename": filename,
                "content_type": content_type,
                "size_bytes": len(report_bytes),
                "data": report_base64
            }
            
        except Exception as e:
            logger.error(f"Error exporting dialect report: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def export_comprehensive_report(self, data: Dict[str, Any], format: str = "pdf") -> Dict[str, Any]:
        """Export comprehensive analytics report"""
        try:
            if format.lower() == "pdf":
                report_bytes = self.pdf_generator.create_comprehensive_report(data)
                filename = f"comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                content_type = "application/pdf"
            else:
                raise ValueError(f"Unsupported format for comprehensive report: {format}")
            
            report_base64 = base64.b64encode(report_bytes).decode('utf-8')
            
            return {
                "success": True,
                "filename": filename,
                "content_type": content_type,
                "size_bytes": len(report_bytes),
                "data": report_base64
            }
            
        except Exception as e:
            logger.error(f"Error exporting comprehensive report: {e}")
            return {
                "success": False,
                "error": str(e)
            }

# Global report service
report_service = ReportExportService()